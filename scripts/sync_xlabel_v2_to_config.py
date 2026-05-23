#!/usr/bin/env python3
"""Sync xlabel_v2.xlsx into config/recommender/xft module YAML files."""

from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "xlabel_v2.xlsx"
CONFIG_DIR = ROOT / "config" / "recommender" / "xft"
MODULES_DIR = CONFIG_DIR / "modules.d"


def main() -> None:
    wb = load_workbook(XLSX_PATH, data_only=True)
    modules = _rows(wb, "01_模块定义")
    labels = _rows(wb, "02_标签定义")
    indicators = _rows(wb, "03_指标定义")
    local_rules = _rows(wb, "04_本地证据规则")
    web_rules = _rows(wb, "05_Web搜索策略")
    llm_rules = _rows(wb, "06_LLM判断边界")
    kyc_rows = _rows(wb, "07_销售话术KYC")

    labels_by_module = _group(labels, "module_id")
    indicators_by_label = _group(indicators, "module_id", "label_id")
    local_by_indicator = _group(local_rules, "module_id", "label_id", "indicator_id")
    web_by_indicator = {_key(row): row for row in web_rules if row.get("module_id") and row.get("indicator_id")}
    llm_by_indicator = {_key(row): row for row in llm_rules if row.get("module_id") and row.get("indicator_id")}
    kyc_by_label = {
        (row["module_id"], row["label_id"]): row for row in kyc_rows if row.get("module_id") and row.get("label_id")
    }
    context = {
        "indicators_by_label": indicators_by_label,
        "local_by_indicator": local_by_indicator,
        "web_by_indicator": web_by_indicator,
        "llm_by_indicator": llm_by_indicator,
        "kyc_by_label": kyc_by_label,
    }

    MODULES_DIR.mkdir(parents=True, exist_ok=True)
    for module_row in modules:
        module_id = str(module_row["module_id"])
        payload = _module_payload(
            module_row=module_row,
            labels=labels_by_module.get((module_id,), []),
            context=context,
        )
        output_path = MODULES_DIR / f"{module_id}.yaml"
        output_path.write_text(
            yaml.safe_dump(payload, allow_unicode=True, sort_keys=False, width=1000),
            encoding="utf-8",
        )
        sys.stdout.write(f"wrote {output_path.relative_to(ROOT)}\n")


def _rows(wb: Any, sheet_name: str) -> list[dict[str, Any]]:
    ws = wb[sheet_name]
    headers = [_clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        row = {
            headers[col - 1]: _clean(ws.cell(row_idx, col).value)
            for col in range(1, ws.max_column + 1)
            if headers[col - 1]
        }
        if any(value not in ("", None) for value in row.values()):
            rows.append(row)
    return rows


def _clean(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def _group(rows: list[dict[str, Any]], *fields: str) -> dict[tuple[Any, ...], list[dict[str, Any]]]:
    grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = {}
    for row in rows:
        if all(row.get(field) not in ("", None) for field in fields):
            grouped.setdefault(tuple(row[field] for field in fields), []).append(row)
    return grouped


def _key(row: dict[str, Any]) -> tuple[Any, Any, Any]:
    return (row["module_id"], row["label_id"], row["indicator_id"])


def _module_payload(
    *,
    module_row: dict[str, Any],
    labels: list[dict[str, Any]],
    context: dict[str, Any],
) -> dict[str, Any]:
    module_id = module_row["module_id"]
    module = {
        "module_id": module_id,
        "module_name": module_row["module_name"],
        "priority": int(module_row["module_priority"]),
        "base_score": int(module_row["module_priority"]),
        "acceptance_policy": {
            "levels": [
                {
                    "result": "高",
                    "min_matched_labels": 3,
                    "conclusion": module_row["high_acceptance_rule"],
                },
                {
                    "result": "中高",
                    "min_matched_labels": 2,
                    "conclusion": module_row["mid_acceptance_rule"],
                },
                {
                    "result": "低",
                    "min_matched_labels": 0,
                    "conclusion": module_row["low_acceptance_rule"],
                },
            ]
        },
        "labels": [],
        "marketing_points": {},
    }
    for label_row in labels:
        label_id = label_row["label_id"]
        label_payload = {
            "label_id": label_id,
            "label_name": label_id,
            "description": _label_description(label_row),
            "min_matched_indicators": int(label_row["min_matched_indicators"]),
            "indicators": [],
        }
        for indicator_row in context["indicators_by_label"].get((module_id, label_id), []):
            key = _key(indicator_row)
            label_payload["indicators"].append(
                _indicator_payload(
                    indicator_row=indicator_row,
                    local_rules=context["local_by_indicator"].get(key, []),
                    web_rule=context["web_by_indicator"].get(key),
                    llm_rule=context["llm_by_indicator"].get(key),
                )
            )
        module["labels"].append(label_payload)
        kyc = context["kyc_by_label"].get((module_id, label_id))
        if kyc:
            module["marketing_points"][label_id] = {
                "recommendation": kyc["recommendation"],
                "sale_rule": kyc["sale_rule"],
                "kyc_questions": [
                    kyc[key] for key in ("kyc_question_1", "kyc_question_2", "kyc_question_3") if kyc.get(key)
                ],
            }
    return module


def _label_description(row: dict[str, Any]) -> str:
    return (
        f"{row['label_business_scene']} 痛点：{row['customer_pain']} "
        f"切入：{row['sales_entry_point']} 角色：{row['label_role']}。"
    )


def _indicator_payload(
    *,
    indicator_row: dict[str, Any],
    local_rules: list[dict[str, Any]],
    web_rule: dict[str, Any] | None,
    llm_rule: dict[str, Any] | None,
) -> dict[str, Any]:
    evaluator = str(indicator_row["recommended_evaluator"])
    local_possible = any(rule.get("rule_result_when_matched") == "possible" for rule in local_rules)
    if local_possible and evaluator in {"rule", "hybrid"}:
        evaluator = "hybrid"
    payload: dict[str, Any] = {
        "indicator_id": indicator_row["indicator_id"],
        "indicator_name": indicator_row["indicator_id"],
        "evaluator": evaluator,
        "standard": (f"{indicator_row['indicator_business_meaning']} {indicator_row['standard_or_boundary_note']}"),
        "evidence_hints": _evidence_hints(indicator_row, local_rules, web_rule),
    }
    if evaluator in {"llm", "hybrid", "llm_web"}:
        payload["prompt"] = _prompt(indicator_row, llm_rule)
    data_sources = [_data_source(rule) for rule in local_rules]
    if data_sources:
        payload["data_sources"] = data_sources
    if evaluator in {"rule", "hybrid"}:
        payload["rule"] = _rule(local_rules)
    if evaluator == "hybrid":
        payload["merge_policy"] = "llm_confirm" if local_possible else "rule_first"
    if web_rule:
        payload["web_search"] = _web_search(web_rule)
    return payload


def _prompt(indicator_row: dict[str, Any], llm_rule: dict[str, Any] | None) -> str:
    if not llm_rule:
        return (
            f"请判断企业是否满足指标：{indicator_row['indicator_id']}。\n"
            f"标准：{indicator_row['indicator_business_meaning']} "
            f"{indicator_row['standard_or_boundary_note']}\n"
            "证据不足时返回 unknown 或 possible；不得用行业常识直接推断为 matched。"
        )
    return "\n".join(
        [
            f"请判断企业是否满足指标：{indicator_row['indicator_id']}。",
            f"任务：{llm_rule['llm_task']}",
            f"matched：{llm_rule['matched_condition']}",
            f"possible：{llm_rule['possible_condition']}",
            f"not_matched：{llm_rule['not_matched_condition']}",
            f"unknown：{llm_rule['unknown_condition']}",
            f"禁止推断：{llm_rule['must_not_infer']}",
            "只能基于结构化证据、招聘证据或权威公开证据判断，不得编造。",
        ]
    )


def _data_source(rule: dict[str, Any]) -> dict[str, Any]:
    source: dict[str, Any] = {
        "type": rule["evidence_type"],
        "op": rule["operator"],
        "description": "来自 xlabel_v2.xlsx 本地证据规则。",
    }
    if rule["evidence_type"] == "field":
        source["path"] = rule["profile_path"]
    else:
        source["table"] = rule["table_name"]
        source["field"] = rule["table_field"]
        source["limit"] = 50
    if rule.get("keywords"):
        source["keywords"] = _split_keywords(str(rule["keywords"]))
    if rule.get("min_matches") not in ("", None):
        source["min_matches"] = int(rule["min_matches"])
    if rule.get("value") not in ("", None):
        source["value"] = _scalar(rule["value"])
    elif rule.get("threshold") not in ("", None):
        source["value"] = _scalar(rule["threshold"])
    return source


def _rule(local_rules: list[dict[str, Any]]) -> dict[str, Any]:
    if not local_rules:
        return {"source_field": "__xlsx_no_local_rule__", "op": "exists"}
    rule = local_rules[0]
    if rule["evidence_type"] == "field":
        value = rule.get("value") if rule.get("value") not in ("", None) else rule.get("threshold")
        if value in ("", None) and rule.get("keywords"):
            value = _split_keywords(str(rule["keywords"]))
        return {
            "source_field": rule["profile_path"],
            "op": rule["operator"],
            **({"value": _scalar(value)} if value not in ("", None) else {}),
        }
    value = _split_keywords(str(rule["keywords"])) if rule.get("keywords") else None
    return {
        "source_field": "__xlsx_table_only_rule__",
        "op": rule["operator"],
        **({"value": value} if value else {}),
    }


def _web_search(row: dict[str, Any]) -> dict[str, Any]:
    fixed_queries = [row[key] for key in ("fixed_query_1", "fixed_query_2") if row.get(key)]
    auto_enabled = row.get("auto_query_enabled") == "是"
    return {
        "when": row["web_when"],
        "effect": row["web_effect"],
        "fixed_queries": fixed_queries,
        "auto": {
            "enabled": auto_enabled,
            "max_queries": 1 if auto_enabled else 0,
            "intent": row.get("auto_query_intent") or "",
        },
        "max_auto_rounds": 1 if auto_enabled else 0,
        "max_results": 5,
    }


def _evidence_hints(
    indicator_row: dict[str, Any],
    local_rules: list[dict[str, Any]],
    web_rule: dict[str, Any] | None,
) -> list[str]:
    hints: list[str] = []
    for rule in local_rules:
        if rule.get("keywords"):
            hints.extend(_split_keywords(str(rule["keywords"])))
    if web_rule:
        hints.extend(str(web_rule.get(key) or "") for key in ("fixed_query_1", "fixed_query_2"))
        if web_rule.get("auto_query_intent"):
            hints.append(str(web_rule["auto_query_intent"]))
    if not hints:
        hints.extend(_tokens(str(indicator_row["indicator_business_meaning"])))
        hints.extend(_tokens(str(indicator_row["standard_or_boundary_note"])))
    deduped: list[str] = []
    for raw_hint in hints:
        normalized = raw_hint.strip()
        if normalized and normalized not in deduped:
            deduped.append(normalized)
    return deduped[:12]


def _split_keywords(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[/,，、；;\r\n\t]+", value) if item.strip()]


def _tokens(value: str) -> list[str]:
    return re.findall(r"[\u4e00-\u9fffA-Za-z0-9]{2,}", value)


def _scalar(value: Any) -> Any:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        stripped = value.strip()
        if re.fullmatch(r"-?\d+", stripped):
            return int(stripped)
        if re.fullmatch(r"-?\d+\.\d+", stripped):
            return float(stripped)
    return value


if __name__ == "__main__":
    main()
