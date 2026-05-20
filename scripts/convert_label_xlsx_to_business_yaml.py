#!/usr/bin/env python3
"""Convert label.xlsx business specs into a draft business_modules.yaml."""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

MAX_HINTS = 8


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", nargs="?", default="label.xlsx", help="Path to label.xlsx")
    parser.add_argument(
        "-o",
        "--output",
        default="config/recommender/xft/business_modules.generated.yaml",
        help="Draft YAML output path",
    )
    parser.add_argument("--sheet", default="最新版（0204-全）", help="Worksheet name")
    args = parser.parse_args()

    payload = convert(Path(args.input), sheet_name=args.sheet)
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(yaml.safe_dump(payload, allow_unicode=True, sort_keys=False), encoding="utf-8")
    sys.stdout.write(f"wrote {out}\n")


def convert(path: Path, *, sheet_name: str) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        msg = f"empty worksheet: {sheet_name}"
        raise ValueError(msg)
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    modules: dict[str, dict[str, Any]] = {}
    last_module = ""
    last_label = ""
    last_kyc = ""
    for raw in rows[1:]:
        row = dict(zip(headers, raw, strict=False))
        module_name = _cell(row.get("模块名称")) or last_module
        label_name = _cell(row.get("企业属性标签")) or last_label
        kyc = _cell(row.get("Kyc关键点")) or last_kyc
        indicator_name = _cell(row.get("指标名称"))
        data_source = _cell(row.get("企业新知数据源"))
        llm_strategy = _cell(row.get("结合大模型"))
        standard = _cell(row.get("量化标准"))
        acceptance = _cell(row.get("产品接受度判断规则"))
        recommendation = _cell(row.get("推荐理由"))
        simple_recommendation = _cell(row.get("推荐理由（简化）"))
        if not (module_name and label_name and indicator_name):
            continue
        last_module = module_name
        last_label = label_name
        last_kyc = kyc
        module_id = _slug(module_name)
        label_id = _slug(label_name)
        indicator_id = _slug(indicator_name)
        module = modules.setdefault(
            module_id,
            {
                "module_id": module_id,
                "module_name": module_name,
                "priority": 50,
                "base_score": 0,
                "acceptance_policy": _acceptance_policy(acceptance),
                "labels": {},
                "marketing_points": {},
            },
        )
        label = module["labels"].setdefault(
            label_id,
            {
                "label_id": label_id,
                "label_name": label_name,
                "description": "",
                "min_matched_indicators": 1,
                "indicators": [],
            },
        )
        evaluator, web_search = _evaluator(llm_strategy)
        indicator: dict[str, Any] = {
            "indicator_id": indicator_id,
            "indicator_name": indicator_name,
            "evaluator": evaluator,
            "standard": standard or "待根据业务规格补充量化标准。",
            "data_sources": _data_sources(data_source),
            "prompt": f"请判断企业是否满足指标：{indicator_name}。只能基于证据判断，不得编造。",
            "evidence_hints": _hints(indicator_name, standard),
        }
        if web_search:
            indicator["web_search"] = web_search
        if evaluator == "rule" and not indicator["data_sources"]:
            indicator["evaluator"] = "llm"
        label["indicators"].append(indicator)
        if recommendation or simple_recommendation or kyc:
            module["marketing_points"][label_id] = {
                "recommendation": recommendation or simple_recommendation,
                "sale_rule": simple_recommendation or recommendation,
                "kyc_questions": _split_numbered(kyc),
            }
    normalized_modules = []
    for module in modules.values():
        module["labels"] = list(module["labels"].values())
        normalized_modules.append(module)
    return {
        "version": "1.0",
        "scenario": "sales_recommendation",
        "scoring": {
            "indicator_scores": {"matched": 10, "possible": 5, "unknown": 0, "not_matched": 0},
            "label_scores": {"matched": 30, "possible": 15, "unknown": 0, "not_matched": 0},
        },
        "acceptance_policy": _default_acceptance_policy(),
        "modules": normalized_modules,
    }


def _cell(value: Any) -> str:
    return str(value).strip() if value not in (None, "") else ""


def _slug(value: str) -> str:
    text = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "_", value).strip("_").lower()
    return text[:64] or "item"


def _evaluator(strategy: str) -> tuple[str, dict[str, Any] | None]:
    if "官网" in strategy or "新闻" in strategy or "舆情" in strategy:
        return "llm_web", {"fixed_queries": ["{company_name} 官网", "{company_name} 新闻"], "auto": False}
    if "LLM" in strategy.upper() or "大模型" in strategy:
        return "llm", None
    return "rule", None


def _data_sources(raw: str) -> list[dict[str, Any]]:
    text = raw.lower()
    sources: list[dict[str, Any]] = []
    if not raw or raw == "无":
        return sources
    if "recruit" in text or "招聘" in raw:
        sources.append(
            {"type": "table", "table": "recruitments", "field": "title", "op": "text_contains", "keywords": []}
        )
    if "qualification" in text or "资质" in raw:
        sources.append({"type": "table", "table": "qualifications", "field": "qualification_name", "op": "exists"})
    if "branch" in text or "分支" in raw:
        sources.append({"type": "table", "table": "branches", "field": "branch_name", "op": "exists"})
    if "intellectual" in text or "知识产权" in raw or "专利" in raw:
        sources.append({"type": "field", "path": "ip_counts.patent", "op": ">", "value": 0})
    if "label" in text or "标签" in raw:
        sources.append({"type": "field", "path": "labels", "op": "exists"})
    if "query_company" in text or "员工" in raw:
        sources.append({"type": "field", "path": "employee_count", "op": "exists"})
    return sources


def _hints(*parts: str) -> list[str]:
    hints: list[str] = []
    for part in parts:
        for token in re.findall(r"[\u4e00-\u9fffA-Za-z0-9]{2,}", part or ""):
            if token not in hints:
                hints.append(token)
            if len(hints) >= MAX_HINTS:
                return hints
    return hints


def _split_numbered(text: str) -> list[str]:
    if not text:
        return []
    pieces = re.split(r"\s*\d+[、.]\s*", text)
    return [piece.strip() for piece in pieces if piece.strip()]


def _acceptance_policy(text: str) -> dict[str, Any] | None:
    if not text:
        return None
    return _default_acceptance_policy()


def _default_acceptance_policy() -> dict[str, Any]:
    return {
        "levels": [
            {
                "result": "高",
                "min_matched_labels": 3,
                "conclusion": "企业满足{attributes_number}个属性标签及{indicators_number}个指标，接受度为高。",
            },
            {
                "result": "中高",
                "min_matched_labels": 2,
                "conclusion": "企业满足{attributes_number}个属性标签及{indicators_number}个指标，接受度为中高。",
            },
            {
                "result": "低",
                "min_matched_labels": 0,
                "conclusion": "企业满足{attributes_number}个属性标签及{indicators_number}个指标，接受度为低。",
            },
        ]
    }


if __name__ == "__main__":
    main()
