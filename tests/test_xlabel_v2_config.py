from __future__ import annotations

from pathlib import Path

import openpyxl

from xft.pipeline.recommender.config_loader import load_recommendation_config


SCENARIO_DIR = Path("config/recommender/xft")
XLSX_PATH = Path("xlabel_v2.xlsx")


def _indicator_map() -> dict[tuple[str, str, str], object]:
    config = load_recommendation_config(SCENARIO_DIR)
    return {
        (module.module_id, label.label_id, indicator.indicator_id): indicator
        for module in config.modules
        for label in module.labels
        for indicator in label.indicators
    }


def test_xft_config_matches_xlabel_v2_business_shape() -> None:
    config = load_recommendation_config(SCENARIO_DIR)

    labels_by_module = {module.module_id: [label.label_id for label in module.labels] for module in config.modules}

    assert labels_by_module == {
        "假勤管理": ["多规则考勤", "倒班排班", "多地点用工", "合规追溯压力", "科技数字化接受度"],
        "差旅报销": ["售后现场交付", "项目制实施交付", "渠道销售拓展", "海外商务出行"],
        "日常报销": ["日常外勤垫付", "门店网点运营", "研发试制费用", "行政办公分散"],
        "对公报账": ["采购复杂付款", "合同付款管控", "跨区域采购协同", "公开供应商合作"],
        "个税管理": ["多地人员组织", "多法人集团", "治理合规强化", "派驻与灵活用工"],
        "进项发票": ["供应链复杂", "多主体收票", "税务审计风控", "上市披露增强证据"],
        "销项发票": ["高频订单开票", "开票时效压力", "渠道客户分散", "业财税数字化", "多主体开票管控", "政策试点身份"],
    }


def test_xft_config_includes_xlabel_v2_critical_indicators() -> None:
    indicators = _indicator_map()

    expected = {
        ("假勤管理", "合规追溯压力", "劳动争议与工时争议"): "llm_web",
        ("假勤管理", "科技数字化接受度", "科技资质认证"): "rule",
        ("假勤管理", "科技数字化接受度", "知识产权数量门槛"): "hybrid",
        ("进项发票", "上市披露增强证据", "年报供应商集中度"): "llm_web",
        ("销项发票", "开票时效压力", "先票后款或货票同行线索"): "llm_web",
        ("销项发票", "政策试点身份", "全电发票试点证据"): "llm_web",
        ("个税管理", "治理合规强化", "股票代码上市身份"): "rule",
    }
    for key, evaluator in expected.items():
        assert key in indicators
        assert indicators[key].evaluator == evaluator


def test_xft_config_uses_strict_shared_rule_thresholds() -> None:
    indicators = _indicator_map()

    tech = indicators[("假勤管理", "科技数字化接受度", "科技资质认证")]
    assert any(
        source.type == "table"
        and source.table == "qualifications"
        and source.field == "qualification_name"
        and source.op == "contains_any"
        and "高新技术企业" in source.keywords
        for source in tech.data_sources
    )
    assert all(
        not (source.type == "field" and source.path == "labels" and source.op == "exists")
        for source in tech.data_sources
    )

    ip = indicators[("假勤管理", "科技数字化接受度", "知识产权数量门槛")]
    assert ip.rule is not None
    assert ip.rule.source_field == "ip_counts.patent"
    assert ip.rule.op == ">="
    assert ip.rule.value == 3

    stock = indicators[("个税管理", "治理合规强化", "股票代码上市身份")]
    assert stock.rule is not None
    assert stock.rule.source_field == "stock_code"
    assert stock.rule.op == "exists"


def test_xft_config_syncs_v2_runtime_calibration_rules() -> None:
    config = load_recommendation_config(SCENARIO_DIR)
    indicators = _indicator_map()

    attendance = next(module for module in config.modules if module.module_id == "假勤管理")
    multi_rule = next(label for label in attendance.labels if label.label_id == "多规则考勤")
    assert multi_rule.min_matched_indicators == 2

    manufacturing = indicators[("假勤管理", "多规则考勤", "制造或生产型业务")]
    assert manufacturing.evaluator == "hybrid"
    assert manufacturing.merge_policy == "llm_confirm"
    assert manufacturing.rule is not None
    assert manufacturing.rule.source_field == "industry"

    rd = indicators[("假勤管理", "科技数字化接受度", "研发岗位线索")]
    assert any(
        source.type == "table"
        and source.table == "recruitments"
        and source.field == "title"
        and "研发" in source.keywords
        for source in rd.data_sources
    )

    high_freq = indicators[("销项发票", "高频订单开票", "高频交易产品线索")]
    assert high_freq.evaluator == "hybrid"
    assert high_freq.merge_policy == "llm_confirm"
    assert any(
        source.type == "field" and source.path == "business_scope" and {"电商", "经销", "批发"} <= set(source.keywords)
        for source in high_freq.data_sources
    )


def test_xft_config_has_high_intent_web_queries_for_v2_web_only_indicators() -> None:
    indicators = _indicator_map()

    labor = indicators[("假勤管理", "合规追溯压力", "劳动争议与工时争议")]
    assert labor.web_search is not None
    assert labor.web_search.when == "always"
    assert any("劳动争议" in query and "考勤" in query for query in labor.web_search.fixed_queries)

    disclosure = indicators[("进项发票", "上市披露增强证据", "年报供应商集中度")]
    assert disclosure.web_search is not None
    assert any("年报" in query and "供应商" in query for query in disclosure.web_search.fixed_queries)

    e_invoice = indicators[("销项发票", "政策试点身份", "全电发票试点证据")]
    assert e_invoice.web_search is not None
    assert any("全电发票" in query or "数电票" in query for query in e_invoice.web_search.fixed_queries)


def test_xft_config_syncs_updated_xlabel_v2_priorities_and_web_policy() -> None:
    config = load_recommendation_config(SCENARIO_DIR)
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    module_sheet = wb["01_模块定义"]
    priorities = {
        module_sheet.cell(row, 1).value: module_sheet.cell(row, 3).value for row in range(2, module_sheet.max_row + 1)
    }
    assert {module.module_id: module.priority for module in config.modules} == priorities

    web_sheet = wb["05_Web搜索策略"]
    web_headers = {web_sheet.cell(1, col).value: col for col in range(1, web_sheet.max_column + 1)}
    expected_auto_enabled = {
        (
            web_sheet.cell(row, web_headers["module_id"]).value,
            web_sheet.cell(row, web_headers["label_id"]).value,
            web_sheet.cell(row, web_headers["indicator_id"]).value,
        )
        for row in range(2, web_sheet.max_row + 1)
        if web_sheet.cell(row, web_headers["auto_query_enabled"]).value == "是"
    }
    actual_auto_enabled = {
        key
        for key, indicator in _indicator_map().items()
        if indicator.web_search is not None and indicator.web_search.auto.enabled
    }
    assert actual_auto_enabled == expected_auto_enabled
    assert len(actual_auto_enabled) == 4


def test_xft_config_syncs_xlabel_v2_label_specific_kyc_questions() -> None:
    config = load_recommendation_config(SCENARIO_DIR)
    questions = [
        question
        for module in config.modules
        for point in module.marketing_points.values()
        for question in point.kyc_questions
    ]

    assert len(questions) == 31 * 3
    assert len(set(questions)) == len(questions)
    assert "现在人工、Excel，还是已有系统？" not in questions
    assert "每月大概花多少时间？" not in questions

    attendance = next(module for module in config.modules if module.module_id == "假勤管理")
    assert attendance.marketing_points["倒班排班"].kyc_questions == [
        "是否有夜班、两班倒、轮班或临时调班，班次由谁维护？",
        "排班表和打卡记录是否能自动匹配，还是月底人工核对？",
        "调班、顶班、跨天班导致薪资或工时争议的频率高不高？",
    ]
